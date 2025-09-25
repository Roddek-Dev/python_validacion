#!/usr/bin/env python3
"""
Organizador de Documentos - Script de automatización para clasificación de documentos

Este script automatiza la clasificación, organización y renombrado de documentos
personales usando un algoritmo de puntuación basado en múltiples factores:
- Prefijos numéricos en nombres de archivo
- Palabras clave en nombres de archivo y carpetas
- Análisis de contenido (opcional)
- OCR para imágenes y PDFs escaneados (opcional)

Uso:
    python organizador_documentos.py --origen /ruta/origen --destino /ruta/destino
    python organizador_documentos.py --origen /ruta/origen --destino /ruta/destino --dry-run false
    python organizador_documentos.py --origen /ruta/origen --destino /ruta/destino --config config.yml --enable-content --enable-ocr

Argumentos:
    --origen: Ruta a la carpeta de origen desorganizada (obligatorio)
    --destino: Ruta a la carpeta de destino organizada (obligatorio)
    --config: Ruta al archivo config.yml (opcional)
    --mode: Modo de operación: 'copy' o 'move' (defecto: copy)
    --dry-run: Simula el proceso sin mover archivos (defecto: true)
    --threads: Número de hilos para procesamiento (defecto: 4)
    --enable-content: Habilita búsqueda en contenido de archivos
    --enable-ocr: Habilita OCR para imágenes y PDFs escaneados
    --log: Ruta para archivo de log (defecto: proceso.log)
    --csv: Ruta para archivo CSV de resultados (defecto: resultados.csv)
    --verbose: Activa logs detallados

Librerías requeridas:
    pip install PyYAML pdfplumber python-docx openpyxl Pillow pytesseract tqdm

Nota: Para OCR, Tesseract debe estar instalado en el sistema.
"""

import argparse
import csv
import hashlib
import logging
import os
import re
import shutil
import sys
import threading
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set

import yaml
from tqdm import tqdm

# Importaciones opcionales para análisis de contenido
try:
    import pdfplumber

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document

    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from PIL import Image, ImageEnhance, ImageFilter
    import pytesseract

    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

try:
    from pdf2image import convert_from_path

    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False


class DocumentOrganizer:
    """Clase principal para la organización de documentos."""

    def __init__(self, config: Dict):
        """
        Inicializa el organizador con la configuración proporcionada.

        Args:
            config: Diccionario con la configuración del sistema
        """
        self.config = config
        self.processed_hashes: Set[str] = set()
        self.results: List[Dict] = []
        self.lock = threading.Lock()
        self.logger = self._setup_logger()
        self.stats = {
            "total_files": 0,
            "classified": 0,
            "duplicates": 0,
            "pending": 0,
            "errors": 0,
            "ocr_used": 0,
            "content_analysis_used": 0,
        }

    def _setup_logger(self) -> logging.Logger:
        """Configura el sistema de logging mejorado."""
        logger = logging.getLogger("DocumentOrganizer")
        logger.setLevel(
            logging.DEBUG if self.config.get("verbose", False) else logging.INFO
        )

        for handler in logger.handlers[:]:
            logger.removeHandler(handler)

        log_file = self.config.get("log_file", "proceso.log")
        try:
            from logging.handlers import RotatingFileHandler

            file_handler = RotatingFileHandler(
                log_file,
                maxBytes=10 * 1024 * 1024,  # 10MB
                backupCount=3,
                encoding="utf-8",
            )
        except ImportError:
            file_handler = logging.FileHandler(log_file, encoding="utf-8")

        file_handler.setLevel(logging.DEBUG)

        # Configurar handler para consola
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        formatter = logging.Formatter(
            "%(asctime)s - %(levelname)s - [%(threadName)s] - %(funcName)s:%(lineno)d - %(message)s"
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(logging.Formatter("%(levelname)s - %(message)s"))

        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

        return logger

    def normalize_text(self, text: str) -> str:
        """
        Normalización avanzada de texto mejorada.

        Args:
            text: Texto a normalizar

        Returns:
            Texto normalizado
        """
        if not text:
            return ""

        # Convertir a minúsculas
        text = text.lower()

        text = unicodedata.normalize("NFKD", text)
        text = "".join(c for c in text if not unicodedata.combining(c))

        text = re.sub(r"[-._\s$$$$\[\]{}]+", " ", text)

        text = re.sub(r"[^\w\s]", " ", text)

        text = re.sub(r"\s+", " ", text)

        return text.strip()

    def calculate_file_hash(self, file_path: Path) -> str:
        """
        Calcula el hash SHA256 de un archivo.

        Args:
            file_path: Ruta al archivo

        Returns:
            Hash SHA256 del archivo
        """
        hash_sha256 = hashlib.sha256()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_sha256.update(chunk)
            return hash_sha256.hexdigest()
        except Exception as e:
            self.logger.error(f"Error calculando hash para {file_path}: {e}")
            return ""

    def is_pdf_scanned(self, file_path: Path) -> bool:
        """Detecta si un PDF es escaneado y necesita OCR."""
        if not PDF_AVAILABLE:
            return True  # Asumir escaneado si no hay pdfplumber

        try:
            text_content = ""
            with pdfplumber.open(file_path) as pdf:
                # Analizar primeras 2 páginas
                for i, page in enumerate(pdf.pages[:2]):
                    page_text = page.extract_text()
                    if page_text:
                        text_content += page_text

            # Criterios para detectar PDF escaneado
            text_len = len(text_content.strip())
            if text_len < self.config.get("scanned_pdf_threshold", 100):
                return True

            # Verificar si el texto es mayoritariamente símbolos extraños
            weird_chars = sum(1 for c in text_content if ord(c) > 127 or c in "□■▪▫")
            weird_ratio = weird_chars / max(text_len, 1)

            if weird_ratio > 0.3:  # Más del 30% caracteres extraños
                return True

            # Verificar densidad de espacios (PDFs escaneados mal procesados)
            space_ratio = text_content.count(" ") / max(text_len, 1)
            if space_ratio > 0.7:  # Más del 70% espacios
                return True

            return False

        except Exception as e:
            self.logger.warning(f"Error verificando PDF escaneado {file_path}: {e}")
            return True  # En caso de error, asumir escaneado

    def preprocess_image_for_ocr(self, image: Image.Image) -> Image.Image:
        """Pre-procesa imagen para mejorar OCR."""
        try:
            # Convertir a escala de grises
            if image.mode != "L":
                image = image.convert("L")

            # Mejorar contraste
            enhancer = ImageEnhance.Contrast(image)
            image = enhancer.enhance(2.0)

            # Aplicar filtro de nitidez
            image = image.filter(ImageFilter.SHARPEN)

            # Redimensionar si es muy pequeña (mínimo 300 DPI equivalente)
            width, height = image.size
            if width < 1200 or height < 1200:
                scale_factor = max(1200 / width, 1200 / height)
                new_size = (int(width * scale_factor), int(height * scale_factor))
                image = image.resize(new_size, Image.Resampling.LANCZOS)

            return image

        except Exception as e:
            self.logger.error(f"Error pre-procesando imagen: {e}")
            return image

    def extract_text_from_pdf(self, file_path: Path, max_pages: int = 3) -> str:
        """
        Extracción mejorada de texto de PDF.

        Args:
            file_path: Ruta al archivo PDF
            max_pages: Número máximo de páginas a leer

        Returns:
            Texto extraído del PDF
        """
        if not PDF_AVAILABLE:
            return ""

        try:
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for i, page in enumerate(pdf.pages[:max_pages]):
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"

                        tables = page.extract_tables()
                        for table in tables:
                            for row in table:
                                if row:
                                    text += (
                                        " ".join(str(cell) for cell in row if cell)
                                        + "\n"
                                    )

                    except Exception as page_error:
                        self.logger.warning(
                            f"Error procesando página {i+1} de {file_path}: {page_error}"
                        )
                        continue

            return text

        except Exception as e:
            self.logger.error(f"Error extrayendo texto de PDF {file_path}: {e}")
            return ""

    def extract_text_from_docx(self, file_path: Path) -> str:
        """
        Extrae texto de un archivo DOCX.

        Args:
            file_path: Ruta al archivo DOCX

        Returns:
            Texto extraído del DOCX
        """
        if not DOCX_AVAILABLE:
            return ""

        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            self.logger.error(f"Error extrayendo texto de DOCX {file_path}: {e}")
            return ""

    def extract_text_from_xlsx(self, file_path: Path, max_sheets: int = 1) -> str:
        """
        Extrae texto de un archivo XLSX.

        Args:
            file_path: Ruta al archivo XLSX
            max_sheets: Número máximo de hojas a leer

        Returns:
            Texto extraído del XLSX
        """
        if not XLSX_AVAILABLE:
            return ""

        try:
            workbook = load_workbook(file_path, data_only=True)
            text = ""
            for i, sheet_name in enumerate(workbook.sheetnames[:max_sheets]):
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text += str(cell) + " "
                text += "\n"
            return text
        except Exception as e:
            self.logger.error(f"Error extrayendo texto de XLSX {file_path}: {e}")
            return ""

    def extract_text_from_txt(self, file_path: Path) -> str:
        """
        Extrae texto de un archivo TXT.

        Args:
            file_path: Ruta al archivo TXT

        Returns:
            Texto extraído del TXT
        """
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception as e:
            self.logger.error(f"Error extrayendo texto de TXT {file_path}: {e}")
            return ""

    def extract_text_with_ocr(self, file_path: Path) -> str:
        """
        OCR mejorado con pre-procesamiento.

        Args:
            file_path: Ruta al archivo

        Returns:
            Texto extraído con OCR
        """
        if not OCR_AVAILABLE:
            self.logger.warning("OCR no disponible - instale Pillow y pytesseract")
            return ""

        try:
            extension = file_path.suffix.lower()

            if extension == ".pdf":
                if not self.is_pdf_scanned(file_path):
                    # Si no es escaneado, usar extracción normal
                    return self.extract_text_from_pdf(file_path)

                # Es escaneado, usar OCR
                if not PDF2IMAGE_AVAILABLE:
                    self.logger.warning("pdf2image no disponible para OCR en PDFs")
                    return ""

                try:
                    pages = convert_from_path(
                        str(file_path),
                        dpi=300,  # Alta resolución
                        first_page=1,
                        last_page=min(3, self.config.get("ocr_max_pages", 3)),
                    )

                    text_parts = []
                    ocr_config = "--oem 3 --psm 6 -l " + self.config.get(
                        "ocr_languages", "spa"
                    )

                    for page_img in pages:
                        processed_img = self.preprocess_image_for_ocr(page_img)

                        # Extraer texto con configuración optimizada
                        page_text = pytesseract.image_to_string(
                            processed_img, config=ocr_config
                        )
                        if page_text.strip():
                            text_parts.append(page_text)

                    result = "\n".join(text_parts)
                    if result.strip():
                        self.stats["ocr_used"] += 1
                        self.logger.debug(
                            f"OCR exitoso en PDF: {file_path.name} ({len(result)} chars)"
                        )

                    return result

                except Exception as conv_error:
                    self.logger.error(
                        f"Error convirtiendo PDF para OCR {file_path}: {conv_error}"
                    )
                    return ""

            elif extension in [".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp"]:
                image = Image.open(file_path)
                processed_image = self.preprocess_image_for_ocr(image)

                ocr_config = "--oem 3 --psm 6 -l " + self.config.get(
                    "ocr_languages", "spa"
                )
                text = pytesseract.image_to_string(processed_image, config=ocr_config)

                if text.strip():
                    self.stats["ocr_used"] += 1
                    self.logger.debug(
                        f"OCR exitoso en imagen: {file_path.name} ({len(text)} chars)"
                    )

                return text

        except Exception as e:
            self.logger.error(f"Error en OCR mejorado para {file_path}: {e}")

        return ""

    def extract_content_text(self, file_path: Path) -> str:
        """
        Extrae texto del contenido de un archivo según su extensión.

        Args:
            file_path: Ruta al archivo

        Returns:
            Texto extraído del archivo
        """
        extension = file_path.suffix.lower()

        if extension == ".pdf":
            return self.extract_text_from_pdf(
                file_path,
                self.config.get("pdf_pages_to_read", 3),  # Usar valor de config
            )
        elif extension == ".docx":
            return self.extract_text_from_docx(file_path)
        elif extension == ".xlsx":
            return self.extract_text_from_xlsx(file_path)
        elif extension == ".txt":
            return self.extract_text_from_txt(file_path)

        return ""

    def get_keyword_score(self, keyword: str, category_id: str) -> Tuple[int, str]:
        """Determina la puntuación de una keyword basada en su especificidad."""
        # Keywords específicas (únicas para una categoría)
        specific_keywords = {
            "00": ["check list", "lista chequeo", "formato check list"],
            "01": ["requisicion personal", "solicitud personal", "requisicion"],
            "02": ["hoja vida", "curriculum vitae", "cv profesional"],
            "03": [
                "diploma",
                "acta grado",
                "titulo profesional",
                "certificado estudios",
            ],
            "04": [
                "certificacion laboral",
                "experiencia laboral",
                "constancia laboral",
            ],
            "05": ["afiliacion eps", "salud eps"],
            "06": ["afiliacion arl", "riesgos laborales"],
            "07": ["caja compensacion", "ccf"],
            "08": ["fondo pension", "afiliacion pension"],
            "09": ["contrato trabajo", "contrato laboral"],
            "10": ["cedula ciudadania", "documento identidad"],
            "11": ["examen medico ingreso", "aptitud medica"],
            "12": ["pgn", "procuraduria", "antecedentes disciplinarios"],
            "13": ["cgr", "contraloria", "antecedentes fiscales"],
            "14": ["ponal", "antecedentes judiciales"],
            "15": ["medidas correctivas ponal"],
            "16": ["certificacion bancaria", "cuenta bancaria"],
            "17": ["licencia conduccion"],
            "18": ["tarjeta profesional", "matricula profesional"],
            "19": ["induccion corporativa", "constancia induccion"],
            "20": ["foto"],
            "21": ["tratamiento datos personales", "habeas data", "autorizacion datos"],
            "22": ["inhabilidad incompatibilidad", "declaracion juramentada"],
            "23": ["otros documentos ingreso"],
            "24": ["examen medico periodico", "reubicacion medica"],
            "25": ["permisos", "licencia no remunerada"],
            "27": ["incapacidad medica"],
            "28": ["vacaciones", "solicitud vacaciones"],
            "29": ["encargo"],
            "30": ["cesantias"],
            "31": ["certificaciones laborales"],
            "32": ["deduccion retefuente", "retencion fuente"],
            "33": ["ingresos retenciones"],
            "34": ["verificacion titulo", "poligrafo"],
            "35": ["reinducciones"],
            "36": ["carnet vacunas", "vacunacion"],
            "37": ["paz salvo", "acta retiro", "examen egreso"],
            "38": ["acta entrega puesto"],
            "39": ["validacion"],
            "40": ["proceso disciplinario", "memorando disciplinario"],
            "41": ["otros documentos"],
            "42": ["evaluacion desempeno", "evaluacion rendimiento"],
        }

        # Keywords genéricas (aparecen en múltiples categorías)
        generic_keywords = [
            "carlos",
            "bautista",
            "andres",
            "gonzalez",  # Nombres propios
            "certificado",
            "certificacion",
            "formato",
            "documento",
            "ingreso",
            "laboral",
            "personal",
        ]

        normalized_keyword = self.normalize_text(keyword)

        # Verificar si es específica para esta categoría
        if category_id in specific_keywords:
            for specific in specific_keywords[category_id]:
                if self.normalize_text(specific) == normalized_keyword:
                    return 20, "específica"  # Puntuación alta para keywords específicas

        # Verificar si es genérica
        for generic in generic_keywords:
            if self.normalize_text(generic) in normalized_keyword:
                return 5, "genérica"  # Puntuación baja para keywords genéricas

        # Keyword normal
        return 10, "normal"

    def classify_file(self, file_path: Path) -> Tuple[str, str, int, List[str]]:
        """
        Clasificación mejorada con sistema de puntuación jerárquico.

        Args:
            file_path: Ruta al archivo a clasificar

        Returns:
            Tupla con (categoría, razón, puntuación, palabras_clave_encontradas)
        """
        filename = file_path.name
        normalized_filename = self.normalize_text(filename)
        categories = self.config["categories"]
        keywords = self.config["keywords"]

        scoring_details = {
            "filename_matches": {},
            "folder_matches": {},
            "content_matches": {},
            "ocr_matches": {},
            "penalties": {},
        }

        # Puntuaciones por categoría
        scores = {cat_id: 0 for cat_id in categories.keys()}
        found_keywords = []

        # Clasificación basada en palabras clave y contenido

        for cat_id, cat_keywords in keywords.items():
            filename_score = 0
            filename_matches = []

            for keyword in cat_keywords:
                normalized_keyword = self.normalize_text(keyword)
                if normalized_keyword in normalized_filename:
                    keyword_score, keyword_type = self.get_keyword_score(
                        keyword, cat_id
                    )
                    filename_score += keyword_score
                    filename_matches.append(f"{keyword}({keyword_type})")
                    found_keywords.append(f"filename:{keyword}")

            scores[cat_id] += filename_score
            if filename_matches:
                scoring_details["filename_matches"][cat_id] = filename_matches

        parent_folder = file_path.parent.name
        normalized_parent = self.normalize_text(parent_folder)

        for cat_id, cat_keywords in keywords.items():
            folder_score = 0
            folder_matches = []

            for keyword in cat_keywords:
                normalized_keyword = self.normalize_text(keyword)
                if normalized_keyword in normalized_parent:
                    keyword_score, keyword_type = self.get_keyword_score(
                        keyword, cat_id
                    )
                    # Reducir puntuación para carpeta vs filename
                    folder_score += int(keyword_score * 0.7)
                    folder_matches.append(f"{keyword}({keyword_type})")
                    found_keywords.append(f"folder:{keyword}")

            scores[cat_id] += folder_score
            if folder_matches:
                scoring_details["folder_matches"][cat_id] = folder_matches

        # Paso 4: Regla de Prioridad Baja - Análisis de Contenido
        if self.config.get(
            "enable_content_search", False
        ) and file_path.suffix.lower() in self.config.get(
            "extensions_for_content_search", []
        ):

            content_text = self.extract_content_text(file_path)
            if content_text:
                self.stats["content_analysis_used"] += 1
                normalized_content = self.normalize_text(content_text)

                for cat_id, cat_keywords in keywords.items():
                    content_score = 0
                    content_matches = []

                    for keyword in cat_keywords:
                        normalized_keyword = self.normalize_text(keyword)
                        if normalized_keyword in normalized_content:
                            keyword_score, keyword_type = self.get_keyword_score(
                                keyword, cat_id
                            )
                            content_score += max(1, int(keyword_score * 0.3))
                            content_matches.append(f"{keyword}({keyword_type})")
                            found_keywords.append(f"content:{keyword}")

                    scores[cat_id] += content_score
                    if content_matches:
                        scoring_details["content_matches"][cat_id] = content_matches

        # Paso 5: OCR (si está habilitado)
        if self.config.get("enable_ocr", False):
            ocr_text = self.extract_text_with_ocr(file_path)

            if ocr_text.strip():
                normalized_ocr = self.normalize_text(ocr_text)

                for cat_id, cat_keywords in keywords.items():
                    ocr_score = 0
                    ocr_matches = []

                    for keyword in cat_keywords:
                        normalized_keyword = self.normalize_text(keyword)
                        if normalized_keyword in normalized_ocr:
                            keyword_score, keyword_type = self.get_keyword_score(
                                keyword, cat_id
                            )
                            ocr_score += int(keyword_score * 0.6)
                            ocr_matches.append(f"{keyword}({keyword_type})")
                            found_keywords.append(f"ocr:{keyword}")

                    scores[cat_id] += ocr_score
                    if ocr_matches:
                        scoring_details["ocr_matches"][cat_id] = ocr_matches

        max_score = max(scores.values()) if scores.values() else 0
        confidence_threshold = self.config.get(
            "confidence_threshold", 25
        )  # Umbral más alto

        if max_score >= confidence_threshold:
            # Encontrar categorías con puntuación máxima
            top_categories = [
                cat_id for cat_id, score in scores.items() if score == max_score
            ]

            competing_categories = [
                cat_id
                for cat_id, score in scores.items()
                if score >= max_score * 0.7 and score > 15
            ]

            if len(competing_categories) > 1:
                penalty = len(competing_categories) * 3
                self.logger.warning(
                    f"Penalización por ambigüedad: -{penalty} puntos para {filename}"
                )
                max_score -= penalty
                scoring_details["penalties"]["ambiguity"] = penalty

            # Si después de penalizaciones sigue siendo válido
            if max_score >= confidence_threshold and len(top_categories) == 1:
                category = top_categories[0]
                reason = self._determine_reason(scores[category], found_keywords)
                return category, reason, max_score, found_keywords
            elif len(top_categories) > 1:
                self.logger.warning(
                    f"Empate después de penalizaciones en {top_categories} para: {filename}"
                )
                return (
                    "Pendientes_Revisar",
                    "empate_post_penalizacion",
                    max_score,
                    found_keywords,
                )

        reason = "baja_confianza" if max_score > 0 else "sin_coincidencias"
        self.logger.warning(
            f"Archivo sin clasificar ({max_score} < {confidence_threshold}): {filename}"
        )
        return "Pendientes_Revisar", reason, max_score, found_keywords

    def _determine_reason(self, score: int, found_keywords: List[str]) -> str:
        """
        Determina la razón de la clasificación basada en la puntuación y palabras clave.

        Args:
            score: Puntuación total
            found_keywords: Lista de palabras clave encontradas

        Returns:
            Razón de la clasificación
        """
        if any(kw.startswith("folder:") for kw in found_keywords):
            return "carpeta_padre"
        elif any(kw.startswith("content:") for kw in found_keywords):
            return "contenido"
        elif any(kw.startswith("ocr:") for kw in found_keywords):
            return "ocr"
        else:
            return "nombre_archivo"

    def clean_filename(self, filename: str) -> str:
        """
        Limpia el nombre de archivo eliminando caracteres inválidos en Windows.

        Args:
            filename: Nombre original del archivo

        Returns:
            Nombre de archivo limpio
        """
        # Caracteres inválidos en Windows
        invalid_chars = '<>:"/\\|?*'

        # Reemplazar caracteres inválidos
        clean_name = filename
        for char in invalid_chars:
            clean_name = clean_name.replace(char, "_")

        # Truncar si es demasiado largo
        max_length = self.config.get("max_filename_length", 100)
        if len(clean_name) > max_length:
            name_part, ext = os.path.splitext(clean_name)
            clean_name = name_part[: max_length - len(ext)] + ext

        return clean_name

    def generate_new_filename(
        self, original_path: Path, category_id: str, dest_folder: Path
    ) -> str:
        """
        Genera el nuevo nombre de archivo según las reglas de renombrado.

        Args:
            original_path: Ruta original del archivo
            category_id: ID de la categoría asignada
            dest_folder: Carpeta de destino

        Returns:
            Nuevo nombre de archivo
        """
        category_name = self.config["categories"][category_id]
        original_name = original_path.stem
        extension = original_path.suffix

        # Limpiar nombre original
        clean_original = self.clean_filename(original_name)

        # Verificar si es categoría de archivo único
        single_file_categories = self.config.get("single_file_name_categories", {})

        # Contar archivos que serán movidos a esta categoría
        if (
            category_id in single_file_categories
            and single_file_categories[category_id]
        ):
            # Formato simplificado para archivo único
            new_name = f"{category_name}{extension}"
        else:
            # Formato estándar
            new_name = f"{category_name} - {clean_original}{extension}"

        # Verificar si ya existe y añadir sufijo numérico si es necesario
        final_path = dest_folder / new_name
        counter = 1

        while final_path.exists():
            name_part, ext = os.path.splitext(new_name)
            new_name = f"{name_part} ({counter}){ext}"
            final_path = dest_folder / new_name
            counter += 1

        return new_name

    def process_file(
        self, file_path: Path, origen_path: Path, destino_path: Path
    ) -> Dict:
        """
        Procesa un archivo individual.

        Args:
            file_path: Ruta al archivo a procesar
            origen_path: Ruta de la carpeta origen
            destino_path: Ruta de la carpeta destino

        Returns:
            Diccionario con los resultados del procesamiento
        """
        result = {
            "timestamp": datetime.now().isoformat(),
            "ruta_original": str(file_path.relative_to(origen_path)),
            "hash_sha256": "",
            "categoria_asignada": "",
            "razon_decision": "",
            "puntuacion": 0,
            "palabras_clave_encontradas": "",
            "ruta_destino": "",
            "estado": "",
            "mensaje_error": "",
        }

        try:
            # Calcular hash del archivo
            file_hash = self.calculate_file_hash(file_path)
            result["hash_sha256"] = file_hash

            # Verificar si es duplicado
            with self.lock:
                if file_hash in self.processed_hashes:
                    result["estado"] = "Duplicado"
                    result["categoria_asignada"] = "Duplicados"
                    self.stats["duplicates"] += 1

                    # Mover a carpeta de duplicados
                    duplicados_folder = destino_path / "Duplicados"
                    duplicados_folder.mkdir(exist_ok=True)

                    new_filename = file_path.name
                    dest_path = duplicados_folder / new_filename

                    # Evitar sobrescribir duplicados
                    counter = 1
                    while dest_path.exists():
                        name_part, ext = os.path.splitext(new_filename)
                        new_filename = f"{name_part}_dup_{counter}{ext}"
                        dest_path = duplicados_folder / new_filename
                        counter += 1

                    result["ruta_destino"] = str(dest_path.relative_to(destino_path))

                    if not self.config.get("dry_run", True):
                        if self.config.get("default_mode", "copy") == "move":
                            shutil.move(str(file_path), str(dest_path))
                        else:
                            shutil.copy2(str(file_path), str(dest_path))

                    self.logger.info(f"Archivo duplicado: {file_path.name}")
                    return result

                self.processed_hashes.add(file_hash)

            # Clasificar archivo
            category_id, reason, score, keywords = self.classify_file(file_path)

            result["categoria_asignada"] = category_id
            result["razon_decision"] = reason
            result["puntuacion"] = score
            result["palabras_clave_encontradas"] = "; ".join(keywords)

            if category_id == "Pendientes_Revisar":
                dest_folder = destino_path / "Pendientes_Revisar"
                dest_folder.mkdir(parents=True, exist_ok=True)
                new_filename = file_path.name
                dest_path = dest_folder / new_filename
                self.stats["pending"] += 1

                # Evitar sobrescribir archivos con el mismo nombre
                counter = 1
                while dest_path.exists():
                    name_part, ext = os.path.splitext(file_path.name)
                    new_filename = f"{name_part}_pendiente_{counter}{ext}"
                    dest_path = dest_folder / new_filename
                    counter += 1
            else:
                category_name = self.config["categories"][category_id]
                dest_folder = destino_path / category_name
                dest_folder.mkdir(parents=True, exist_ok=True)

                # Generar nuevo nombre de archivo
                new_filename = self.generate_new_filename(
                    file_path, category_id, dest_folder
                )
                dest_path = dest_folder / new_filename
                self.stats["classified"] += 1

            result["ruta_destino"] = str(dest_path.relative_to(destino_path))
            result["estado"] = "Clasificado"

            # Mover o copiar archivo
            if not self.config.get("dry_run", True):
                if self.config.get("default_mode", "copy") == "move":
                    shutil.move(str(file_path), str(dest_path))
                else:
                    shutil.copy2(str(file_path), str(dest_path))

            if category_id == "Pendientes_Revisar":
                self.logger.warning(
                    f"Archivo pendiente de revisar: {file_path.name} -> {reason}"
                )
            else:
                self.logger.info(
                    f"Archivo procesado: {file_path.name} -> {category_id} ({reason})"
                )

        except Exception as e:
            result["estado"] = "Error"
            result["mensaje_error"] = str(e)
            self.logger.error(f"Error procesando {file_path}: {e}")
            self.stats["errors"] += 1

        return result

    def find_files(self, origen_path: Path) -> List[Path]:
        """
        Encuentra todos los archivos en la carpeta origen y subcarpetas.

        Args:
            origen_path: Ruta de la carpeta origen

        Returns:
            Lista de rutas de archivos encontrados
        """
        files = []
        supported_extensions = {
            ".pdf",
            ".docx",
            ".xlsx",
            ".jpg",
            ".jpeg",
            ".png",
            ".txt",
            ".tiff",
            ".bmp",
        }

        for file_path in origen_path.rglob("*"):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                files.append(file_path)

        return files

    def organize_documents(self, origen_path: Path, destino_path: Path):
        """
        Organiza todos los documentos de la carpeta origen.

        Args:
            origen_path: Ruta de la carpeta origen
            destino_path: Ruta de la carpeta destino
        """
        self.logger.info(f"Iniciando organización de documentos")
        self.logger.info(f"Origen: {origen_path}")
        self.logger.info(f"Destino: {destino_path}")
        self.logger.info(
            f"Modo: {'DRY RUN' if self.config.get('dry_run', True) else self.config.get('default_mode', 'copy').upper()}"
        )

        # Crear carpeta destino
        destino_path.mkdir(parents=True, exist_ok=True)

        # Encontrar todos los archivos
        files = self.find_files(origen_path)
        self.stats["total_files"] = len(files)
        self.logger.info(f"Encontrados {len(files)} archivos para procesar")

        if not files:
            self.logger.warning("No se encontraron archivos para procesar")
            return

        # Procesar archivos con threading
        num_threads = self.config.get("num_threads", 4)

        with ThreadPoolExecutor(max_workers=num_threads) as executor:
            # Crear barra de progreso
            with tqdm(total=len(files), desc="Procesando archivos") as pbar:
                # Enviar tareas
                future_to_file = {
                    executor.submit(
                        self.process_file, file_path, origen_path, destino_path
                    ): file_path
                    for file_path in files
                }

                # Recoger resultados
                for future in as_completed(future_to_file):
                    result = future.result()
                    with self.lock:
                        self.results.append(result)
                    pbar.update(1)

        # Guardar resultados en CSV
        self.save_results_csv()

        # Mostrar resumen
        self.print_summary()

    def save_results_csv(self):
        """Guarda los resultados en un archivo CSV."""
        csv_file = self.config.get("csv_file", "resultados.csv")

        try:
            with open(csv_file, "w", newline="", encoding="utf-8") as f:
                if self.results:
                    writer = csv.DictWriter(f, fieldnames=self.results[0].keys())
                    writer.writeheader()
                    writer.writerows(self.results)

            self.logger.info(f"Resultados guardados en: {csv_file}")
        except Exception as e:
            self.logger.error(f"Error guardando CSV: {e}")

    def print_summary(self):
        """Imprime un resumen del procesamiento mejorado."""
        if not self.results:
            return

        total_files = len(self.results)

        # Contar por estado
        estados = {}
        for result in self.results:
            estado = result["estado"]
            estados[estado] = estados.get(estado, 0) + 1

        # Contar por categoría
        categorias = {}
        pendientes_por_razon = {}
        for result in self.results:
            if result["estado"] == "Clasificado":
                categoria = result["categoria_asignada"]
                categorias[categoria] = categorias.get(categoria, 0) + 1

                if categoria == "Pendientes_Revisar":
                    razon = result["razon_decision"]
                    pendientes_por_razon[razon] = pendientes_por_razon.get(razon, 0) + 1

        print("\n" + "=" * 60)
        print("RESUMEN DE PROCESAMIENTO")
        print("=" * 60)
        print(f"Total de archivos procesados: {total_files}")
        print(
            f"Modo de ejecución: {'DRY RUN (simulación)' if self.config.get('dry_run', True) else 'REAL'}"
        )

        print("\nPor estado:")
        for estado, count in estados.items():
            print(f"  {estado}: {count}")

        if categorias:
            print("\nPor categoría:")
            for categoria, count in sorted(categorias.items()):
                if categoria == "Pendientes_Revisar":
                    print(f"  {categoria}: {count}")
                else:
                    category_name = self.config["categories"].get(categoria, categoria)
                    print(f"  {category_name}: {count}")

        if pendientes_por_razon:
            print("\nArchivos pendientes de revisar por razón:")
            for razon, count in pendientes_por_razon.items():
                razon_desc = {
                    "baja_confianza": "Baja confianza (puntuación insuficiente)",
                    "sin_coincidencias": "Sin coincidencias (no se encontraron keywords)",
                    "empate_post_penalizacion": "Empate después de penalizaciones",
                    "sin_clasificar": "Sin clasificar (puntuación baja)",
                    "ambiguo": "Ambiguo (empate en categorías)",
                }.get(razon, razon)
                print(f"  {razon_desc}: {count}")

        duplicados = estados.get("Duplicado", 0)
        pendientes = categorias.get("Pendientes_Revisar", 0)
        errores = estados.get("Error", 0)

        print(f"\nArchivos duplicados encontrados: {duplicados}")
        print(f"Archivos pendientes de revisar: {pendientes}")
        print(f"Archivos con errores: {errores}")

        print(f"\nEstadísticas de procesamiento:")
        print(
            f"  Análisis de contenido usado: {self.stats['content_analysis_used']} archivos"
        )
        print(f"  OCR usado: {self.stats['ocr_used']} archivos")
        print("=" * 60)


def load_config(config_path: Optional[str] = None) -> Dict:
    """
    Carga la configuración desde archivo YAML o usa valores por defecto.

    Args:
        config_path: Ruta al archivo de configuración

    Returns:
        Diccionario con la configuración
    """
    default_config = {
        "dry_run": True,
        "default_mode": "copy",
        "num_threads": 4,
        "confidence_threshold": 25,  # Umbral más alto para mejor precisión
        "enable_content_search": False,
        "pdf_pages_to_read": 3,  # Leer más páginas para mejor análisis
        "extensions_for_content_search": [".pdf", ".docx", ".txt", ".xlsx"],
        "enable_ocr": False,
        "ocr_languages": "spa",
        "ocr_max_pages": 3,  # Nuevo parámetro para OCR
        "scanned_pdf_threshold": 100,  # Umbral para detectar PDFs escaneados
        "ocr_min_text_threshold": 50,
        "max_filename_length": 100,
        "single_file_name_categories": {"00": True, "02": True, "10": True, "20": True},
        "categories": {
            "00": "00 Formato Check List",
            "01": "01 Requisición del Personal",
            "02": "02 Formato Hoja de Vida Unica",
            "03": "03 Certificados de Estudios",
            "04": "04 Certificaciones de Experiencia Laboral",
            "05": "05 certificación y Afiliación EPS",
            "06": "06 Afiliación ARL",
            "07": "07 Afiliación Caja de Compensación",
            "08": "08 Afiliación Fondo de Pensiones",
            "09": "09 Contrato de Trabajo",
            "10": "10 Documento de Identidad",
            "11": "11 Examen Médico de Ingreso",
            "12": "12 Antecedentes Disciplinarios PGN",
            "13": "13 Antecedentes Fiscales CGR",
            "14": "14 Antecedentes Judiciales PONAL",
            "15": "15 Medidas Correctivas PONAL",
            "16": "16 Certificación Bancaria",
            "17": "17 Licencia de Conducción",
            "18": "18 Tarjeta Profesional",
            "19": "19 Constancia Inducción Corporativa",
            "20": "20 Foto (png)",
            "21": "21 Autorización Tratamiento Datos Personales",
            "22": "22 Declaración de no Inhabilidad e Incompatibilidad",
            "23": "23 Otros Documentos Solicitados para el Ingreso",
            "24": "24 Examen Medico Periódico y Reubicación",
            "25": "25 Permisos",
            "27": "27 Incapacidades",
            "28": "28 Vacaciones",
            "29": "29 Encargos",
            "30": "30 Cesantías",
            "31": "31 Certificaciones Laborales",
            "32": "32 Formatos Dedución Rte Fte",
            "33": "33 Certificados Ingresos y Retenciones",
            "34": "34 Verificaciones",
            "35": "35 Reinducciones",
            "36": "36 Carnet de Vacunas",
            "37": "37 Protocolo de Retiro y documentos de retiro",
            "38": "38 Actas de Entrega",
            "39": "39 Validaciones",
            "40": "40 Procesos Disciplinarios",
            "41": "41 Otros documentos",
            "42": "42 Evaluaciones de desempeño",
        },
        "keywords": {
            "00": [
                "check list",
                "lista chequeo",
                "formato check list",
                "lista de chequeo",
                "formato lista chequeo",
                "expediente laboral",
                "check list expediente",
                "formato verificacion",
                "lista verificacion",
                "checklist documento",
                "control documentos",
                "formato control",
            ],
            "01": [
                "requisicion",
                "solicitud personal",
                "requisicion personal",
                "formato requisicion",
                "requisicion personal",
                "solicitud empleo",
                "requerimiento personal",
                "solicitud cargo",
                "formato solicitud personal",
                "autorizacion contrato",
                "aprobacion vacante",
                "necesidad personal",
            ],
            "02": [
                "hoja de vida",
                "cv",
                "curriculum",
                "hv",
                "curriculum vitae",
                "formato hoja vida",
                "hoja vida unica",
                "aplicacion empleo",
                "datos personales",
                "informacion personal",
                "formato hv",
                "solicitud empleo",
                "perfil profesional",
                "experiencia laboral",
                "estudios realizados",
                "formato curriculum",
            ],
            "03": [
                "estudios",
                "diploma",
                "acta grado",
                "titulo profesional",
                "tecnologo",
                "tecnico",
                "certificado estudios",
                "acta individual grado",
                "diploma grado",
                "titulo academico",
                "certificacion estudios",
                "constancia estudios",
                "grado academico",
                "certificado academico",
                "diplomado",
                "maestria",
                "magister",
                "especializacion",
                "postgrado",
                "educacion continua",
                "certificado curso",
                "diploma curso",
                "formacion academica",
                "estudios realizados",
            ],
            "04": [
                "experiencia laboral",
                "certificacion laboral",
                "constancia laboral",
                "certificaciones experiencia",
                "certificado experiencia",
                "constancia experiencia",
                "historial laboral",
                "antecedentes laborales",
                "certificacion contratos",
                "prestacion servicios",
                "contratos servicios",
                "experiencia profesional",
                "certificado trabajo",
                "constancia trabajo",
                "hoja de vida laboral",
            ],
            "05": [
                "eps",
                "afiliacion salud",
                "salud eps",
                "afiliacion eps",
                "certificacion eps",
                "afiliacion sistema salud",
                "plan salud",
                "medicina prepagada",
                "novedad reingreso",
                "registro eps",
                "seguridad salud",
                "afiliado eps",
                "certificado afiliacion",
                "sistema salud",
                "plan medico",
                "seguro salud",
            ],
            "06": [
                "arl",
                "riesgos laborales",
                "afiliacion arl",
                "certificado arl",
                "seguro riesgos laborales",
                "positiva seguros",
                "certificado afiliacion arl",
                "riesgos profesionales",
                "seguro laboral",
                "cobertura arl",
                "estado afiliacion arl",
                "certificado cobertura",
                "administradora riesgos laborales",
                "seguro trabajo",
            ],
            "07": [
                "caja compensacion",
                "ccf",
                "compensacion familiar",
                "certificado caja compensacion",
                "afiliacion compensacion familiar",
                "caja compensacion familiar",
                "subsidio familiar",
                "beneficios ccf",
                "compensar",
                "colsubsidio",
                "cafam",
                "certificado afiliacion ccf",
                "subsidio caja compensacion",
                "compensacion familiar",
            ],
            "08": [
                "pensiones",
                "fondo pension",
                "afiliacion pension",
                "fondo pensiones obligatorias",
                "porvenir",
                "afp",
                "certificado pensiones",
                "administradora fondos pensiones",
                "pension obligatoria",
                "sistema pensiones",
                "certificado afiliacion pension",
                "fondo pensiones",
                "aportes pension",
            ],
            "09": [
                "contrato",
                "contrato trabajo",
                "contrato laboral",
                "contrato individual trabajo",
                "termino indefinido",
                "contrato empleo",
                "contrato laboral indefinido",
                "contrato personal",
                "vinculacion laboral",
                "contrato termino fijo",
                "contrato servicios",
                "acuerdo laboral",
                "documento contrato",
                "firma contrato",
                "condiciones laborales",
            ],
            "10": [
                "cedula",
                "documento identidad",
                "cc",
                "cedula ciudadania",
                "cedula ciudadana",
                "identificacion personal",
                "documento identificacion",
                "tarjeta identidad",
                "documento nacional",
                "identidad personal",
                "cedula ampliada",
                "documento oficial",
                "identificacion oficial",
                "cedula colombiana",
                "documento personal",
            ],
            "11": [
                "examen medico",
                "examen ingreso",
                "aptitud medica",
                "examen medico ingreso",
                "certificado medico ocupacional",
                "examen pre ingreso",
                "aptitud laboral",
                "examen medico laboral",
                "certificado aptitud",
                "examen ocupacional",
                "valoracion medica ingreso",
                "concepto aptitud",
                "examen salud ocupacional",
                "certificado salud",
                "examen ingreso laboral",
                "aptitud sin restricciones",
            ],
            "12": [
                "pgn",
                "procuraduria",
                "antecedentes disciplinarios",
                "certificado antecedentes",
                "procuraduria general nacion",
                "antecedentes pgn",
                "certificado disciplinario",
                "sanciones inhabilidades",
                "certificado ordinario",
                "siri",
                "registro sanciones",
                "certificado procuraduria",
                "antecedentes sanciones",
                "inhabilidades vigentes",
                "certificado disciplinario pgn",
            ],
            "13": [
                "cgr",
                "contraloria",
                "antecedentes fiscales",
                "contraloria general republica",
                "certificado fiscal",
                "responsable fiscal",
                "antecedentes cgr",
                "certificado contraloria",
                "sibor",
                "boletin responsables fiscales",
                "responsabilidad fiscal",
                "certificado antecedentes fiscales",
                "contraloria delegada",
                "certificado cgr",
            ],
            "14": [
                "ponal",
                "policia",
                "antecedentes judiciales",
                "policia nacional",
                "antecedentes penales",
                "requerimientos judiciales",
                "certificado judicial",
                "antecedentes policia",
                "asuntos judiciales",
                "certificado ponal",
                "antecedentes policia nacional",
                "consulta antecedentes",
                "certificado antecedentes penales",
                "autoridades judiciales",
            ],
            "15": [
                "medidas correctivas",
                "medidas correctivas ponal",
                "registro nacional medidas correctivas",
                "rnmc",
                "certificado medidas correctivas",
                "policia nacional",
                "codigo seguridad convivencia",
                "medidas correctivas pendientes",
                "certificado ponal medidas",
                "sistema registro medidas",
                "convivencia ciudadana",
                "certificado correctivas",
                "medidas correctivas policia",
            ],
            "16": [
                "certificacion bancaria",
                "cuenta bancaria",
                "certificado bancario",
                "cuenta ahorros",
                "cuenta corriente",
                "formato cambio cuenta",
                "cambio cuenta bancaria",
                "certificacion cuenta",
                "datos bancarios",
                "informacion bancaria",
                "banco scotiabank",
                "banco colpatria",
                "rappipay",
                "rappicuenta",
                "pago sueldo",
                "haberes bancarios",
                "certificado financiero",
            ],
            "17": [
                "licencia conduccion",
                "pase",
                "licencia conducir",
                "licencia transito",
                "pase conduccion",
                "permiso conduccion",
                "tarjeta propiedad",
                "soat",
                "seguro vehiculo",
                "ministerio transporte",
                "categoria conduccion",
                "vigencia licencia",
                "documento conduccion",
                "certificado conduccion",
                "licencia tipo",
            ],
            "18": [
                "tarjeta profesional",
                "matricula profesional",
                "copnia",
                "consejo profesional ingenieria",
                "matricula ingeniero",
                "tarjeta ingeniero",
                "certificado profesional",
                "registro profesional",
                "titulo profesional",
                "licencia profesional",
                "matricula no.",
                "tarjeta colegiatura",
                "certificado ejercicio profesional",
            ],
            "19": [
                "induccion",
                "reinduccion",
                "induccion corporativa",
                "constancia induccion",
                "formato induccion",
                "induccion laboral",
                "capacitacion ingreso",
                "bienvenida empresa",
                "mision vision",
                "estructura organizacional",
                "sistema gestion",
                "sg sst",
                "reglamentos politicas",
                "manuales empresa",
                "procesos gestion calidad",
                "certificado induccion",
                "formacion corporativa",
            ],
            "20": [
                "foto",
                "fotografia",
                "foto documento",
                "foto carnet",
                "foto 3x4",
                "foto fondo blanco",
                "foto reciente",
                "foto tamaño carnet",
                "foto identificacion",
                "foto personal",
                "foto expediente",
                "foto laboral",
                "foto ingreso",
                "imagen documento",
            ],
            "21": [
                "tratamiento datos",
                "habeas data",
                "autorizacion datos",
                "datos personales",
                "autorizacion tratamiento datos",
                "proteccion datos personales",
                "ley 1581 2012",
                "consentimiento datos",
                "politica datos personales",
                "titular informacion",
                "privacidad datos",
                "autorizacion fnd",
                "formato datos personales",
                "proteccion informacion",
            ],
            "22": [
                "inhabilidad",
                "incompatibilidad",
                "declaracion juramentada",
                "certificado inhabilidades",
                "conflicto intereses",
                "declaracion jurada",
                "ley 80 1993",
                "ley 1474 2011",
                "actividad ilicita",
                "certificado incompatibilidad",
                "declaracion legal",
                "certificado conflictos",
                "inhabilidades incompatibilidades",
            ],
            "23": [
                "otros ingreso",
                "otros documentos ingreso",
                "evaluacion periodo prueba",
                "formato evaluacion",
                "periodo prueba",
                "evaluacion desempeño ingreso",
                "evaluacion competencias",
                "desempeño periodo prueba",
                "documentos complementarios",
                "evaluacion inicial",
                "formato periodo prueba",
                "evaluacion ingreso",
                "competencia laboral",
                "valoracion periodo prueba",
                "documentos varios ingreso",
                "evaluacion contratacion",
            ],
            "24": [
                "examen periodico",
                "reubicacion",
                "examen medico periodico",
                "certificado medico periodico",
                "examen ocupacional periodico",
                "aptitud laboral periodica",
                "recomendaciones medicas",
                "vigilancia epidemiologica",
                "examen reubicacion",
                "concepto medico periodico",
                "control medico laboral",
                "examen cambio ocupacion",
                "seguimiento medico ocupacional",
            ],
            "25": [
                "permiso",
                "licencia no remunerada",
                "permisos",
                "formato control novedades",
                "permiso personal",
                "licencia estudio",
                "permiso academico",
                "licencia luto",
                "dia familia",
                "permiso votaciones",
                "calamidad domestica",
                "licencia remunerada",
                "permiso horas",
                "permiso dias",
                "novedades personal",
                "ausencia laboral",
            ],
            "27": [
                "incapacidad",
                "incapacidades",
                "incapacidad medica",
                "certificado incapacidad",
                "licencia medica",
                "incapacidad general",
                "enfermedad general",
                "diagnostico medico",
                "incapacidad laboral",
                "certificado medico",
                "incapacidad eps",
                "incapacidad sura",
                "incapacidad suramericana",
                "ausencia medica",
                "licencia enfermedad",
                "incapacidad temporal",
                "reposo medico",
            ],
            "28": [
                "vacaciones",
                "solicitud vacaciones",
                "formato vacaciones",
                "solicitud disfrute vacaciones",
                "periodo vacaciones",
                "dias vacaciones",
                "disfrute vacaciones",
                "calendario vacaciones",
                "programacion vacaciones",
                "ausencia vacaciones",
                "reemplazo vacaciones",
                "descanso vacaciones",
                "periodo descanso",
                "tiempo vacaciones",
            ],
            "29": [
                "encargo",
                "encargos",
                "encargo cargo",
                "encargo funciones",
                "delegacion funciones",
                "suplencia",
                "encargo temporal",
                "asignacion temporal",
                "encargo administrativo",
                "designacion encargo",
                "comision encargo",
                "encargo laboral",
            ],
            "30": [
                "cesantias",
                "retiro cesantias",
                "solicitud cesantias",
                "liquidacion cesantias",
                "certificado cesantias",
                "fondo cesantias",
                "retiro fondos cesantias",
                "cesantias acumuladas",
                "afiliacion cesantias",
                "administradora cesantias",
                "cesantias porvenir",
                "cesantias proteccion",
            ],
            "31": [
                "certificaciones laborales",
                "constancias laborales",
                "certificado laboral",
                "constancia laboral",
                "certificacion trabajo",
                "certificado empleo",
                "constancia trabajo",
                "certificado salarial",
                "certificacion ingresos",
                "constancia salario",
                "certificado funciones",
                "certificacion antiguedad",
                "certificado vinculacion",
                "constancia laboral fnd",
                "certificado tiempo servicio",
            ],
            "32": [
                "deduccion",
                "retefuente",
                "retencion fuente",
                "deduccion retefuente",
                # Frases compuestas específicas
                "formato deduccion",
                "actualizacion deduccion",
                "declaracion renta",
                "certificado intereses",
                "intereses hipotecarios",
                "credito vivienda",
                "retencion en la fuente",
                "deducciones tributarias",
                # Variaciones y sinónimos
                "formulario deduccion",
                "actualizacion retefuente",
                "certificado deduccion",
                "intereses credito",
                "deduccion renta",
                "formato retencion",
                # Términos específicos del contenido
                "medicina prepagada",
                "aportes voluntarios",
                "dependientes",
                "cuentas afc",
                "aportes pensiones",
                "certificacion intereses",
                # Abreviaturas comunes
                "rte fte",
                "rete fte",
                "deduccion rte fte",
            ],
            "33": ["ingresos y retenciones", "ingresos retenciones"],
            "33": [
                "ingresos y retenciones",
                "ingresos retenciones",
                "certificado ingresos",
                "certificado retenciones",
                "certificacion ingresos",
                # Frases compuestas específicas
                "certificado laboral ingresos",
                "constancia ingresos",
                "declaracion renta",
                "certificado tributario",
                "retencion en la fuente",
                "certificado salarial",
                "comprobante ingresos",
                "certificado financiero",
                # Términos específicos del contenido
                "ingresos anuales",
                "retenciones año",
                "certificado fiscal",
                "documento tributario",
                "informe ingresos",
                "estado financiero",
                # Sinónimos y variaciones
                "certificado economico",
                "constancia salarial",
                "comprobante retencion",
                "documento ingresos",
                "certificado renta",
                "retfuente",
                # Abreviaturas comunes
                "cert ingresos",
                "cert retenciones",
                "ingresos retfuente",
            ],
            "34": [
                "verificacion",
                "poligrafo",
                "verificaciones",
                "verificacion titulo",
                # Frases compuestas específicas
                "verificacion titulos",
                "confirmacion titulo",
                "validacion estudios",
                "verificacion antecedentes",
                "verificacion laboral",
                "check referencia",
                "confirmacion empleo",
                "validacion experiencia",
                # Términos específicos del contenido
                "solicitud confirmacion",
                "egresado verificado",
                "titulo educativo",
                "institucion educativa",
                "verificacion academica",
                # Sinónimos y variaciones
                "chequeo",
                "confirmacion",
                "validacion",
                "certificacion veracidad",
            ],
            "35": [
                "reinducciones",
                "reinduccion",
                "capacitacion recurrente",
                "entrenamiento periodico",
                "actualizacion induccion",
                "refresher training",
                "reentrenamiento",
                "induccion repetida",
            ],
            "36": [
                "vacunas",
                "carnet vacunacion",
                "carnet vacunas",
                "vacunacion",
                "certificado vacunas",
                "esquema vacunacion",
                "carnet salud",
                "vacuna covid",
                "certificado sanitario",
                "inmunizacion",
            ],
            "37": [
                "retiro",
                "paz y salvo",
                "acta retiro",
                "examen egreso",
                "paz salvo",
                "liquidacion contrato",
                "fin contrato",
                "terminacion laboral",
                "desvinculacion",
                "carta retiro",
                "proceso salida",
            ],
            "38": [
                "acta entrega",
                "entrega puesto",
                "actas entrega",
                "inventario entrega",
                "transferencia responsabilidades",
                "acta recepcion",
                "entrega cargo",
                "transicion puesto",
            ],
            "39": [
                "validacion",
                "validaciones",
                "aprobacion",
                "revision formal",
                # Frases compuestas
                "validacion documentos",
                "revision cumplimiento",
                "control calidad",
                "verificacion formal",
                "autorizacion documento",
                # Procesos específicos
                "validacion requisitos",
                "revision requisitos",
                "cumplimiento normativo",
                "conformidad documental",
                "validacion procedimiento",
                # Términos relacionados
                "visto bueno",
                "revision aprobacion",
                "validacion formal",
            ],
            "40": [
                "proceso disciplinario",
                "llamado atencion",
                "memorando",
                "memorando disciplinario",
                "sancion disciplinaria",
                "investigacion disciplinaria",
                "correctivo",
                "medida disciplinaria",
                "falta laboral",
                "amonestacion",
                "proceso sancionatorio",
            ],
            "41": [
                "otros",
                "varios",
                "otros documentos",
                "documentos varios",
                "miscelaneos",
                "genericos",
                "no clasificados",
                "otros archivos",
            ],
            "42": [
                "evaluacion desempeño",
                "evaluacion rendimiento",
                "evaluacion desempeno",
                "calificacion desempeño",
                "review desempeño",
                "valoracion laboral",
                "evaluacion competencias",
                "assessment desempeño",
                "resultados evaluacion",
                "feedback desempeño",
            ],
        },
    }

    # Cargar configuración desde ruta especificada o intentar 'config.yml' en el cwd
    try:
        candidate_path: Optional[Path] = None
        if config_path:
            candidate_path = Path(config_path)
        else:
            local_cfg = Path("config.yml")
            if local_cfg.exists():
                candidate_path = local_cfg
        if candidate_path and candidate_path.exists():
            with open(candidate_path, "r", encoding="utf-8") as f:
                user_config = yaml.safe_load(f) or {}
                # Combinar configuración por defecto con la del usuario
                default_config.update(user_config)
    except Exception as e:
        print(f"Error cargando configuración: {e}")
        print("Usando configuración por defecto")

    return default_config


def main():
    """Función principal del script."""
    parser = argparse.ArgumentParser(
        description="Organizador de Documentos - Automatización de clasificación de documentos",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  python organizador_documentos.py --origen /ruta/origen --destino /ruta/destino
  python organizador_documentos.py --origen /ruta/origen --destino /ruta/destino --dry-run false
  python organizador_documentos.py --origen /ruta/origen --destino /ruta/destino --config config.yml --enable-content --enable-ocr
        """,
    )

    # Valores por defecto para simplificar el comando
    default_origen = str(Path.home() / "Escritorio" / "Documentos_Desordenados")
    default_destino = str(Path.home() / "Escritorio" / "Documentos_Organizados")

    parser.add_argument(
        "--origen", required=False, default=default_origen, help="Ruta a la carpeta de origen desorganizada"
    )
    parser.add_argument(
        "--destino", required=False, default=default_destino, help="Ruta a la carpeta de destino organizada"
    )
    parser.add_argument("--config", help="Ruta al archivo config.yml")
    parser.add_argument(
        "--mode", choices=["copy", "move"], default="copy", help="Modo de operación"
    )
    parser.add_argument(
        "--dry-run",
        type=str,
        default="true",
        choices=["true", "false"],
        help="Simular proceso sin mover archivos",
    )
    parser.add_argument(
        "--threads", type=int, default=4, help="Número de hilos para procesamiento"
    )
    parser.add_argument(
        "--enable-content", action="store_true", help="Habilitar búsqueda en contenido"
    )
    parser.add_argument("--enable-ocr", action="store_true", help="Habilitar OCR")
    parser.add_argument("--log", default="proceso.log", help="Archivo de log")
    parser.add_argument(
        "--csv", default="resultados.csv", help="Archivo CSV de resultados"
    )
    parser.add_argument("--verbose", action="store_true", help="Logs detallados")

    args = parser.parse_args()

    # Validar rutas
    origen_path = Path(args.origen)
    destino_path = Path(args.destino)

    if not origen_path.exists():
        # Crear origen si no existe para una mejor UX
        origen_path.mkdir(parents=True, exist_ok=True)
        print(f"Creada carpeta origen por defecto: {origen_path}")

    if not origen_path.is_dir():
        print(f"Error: La ruta origen no es una carpeta: {origen_path}")
        sys.exit(1)

    # Cargar configuración
    config = load_config(args.config)

    # Actualizar configuración con argumentos de línea de comandos
    config.update(
        {
            "default_mode": args.mode,
            "dry_run": args.dry_run.lower() == "true",
            "num_threads": args.threads,
            "enable_content_search": args.enable_content,
            "enable_ocr": args.enable_ocr,
            "log_file": args.log,
            "csv_file": args.csv,
            "verbose": args.verbose,
        }
    )

    # Verificar dependencias opcionales
    if config["enable_content_search"]:
        missing_deps = []
        if not PDF_AVAILABLE:
            missing_deps.append("pdfplumber")
        if not DOCX_AVAILABLE:
            missing_deps.append("python-docx")
        if not XLSX_AVAILABLE:
            missing_deps.append("openpyxl")

        if missing_deps:
            print(
                f"Error: Para habilitar búsqueda en contenido, instale: {', '.join(missing_deps)}"
            )
            sys.exit(1)

    if config["enable_ocr"] and not OCR_AVAILABLE:
        print("Error: Para habilitar OCR, instale: pip install Pillow pytesseract")
        print("También debe instalar Tesseract en su sistema")
        sys.exit(1)

    # Crear organizador y ejecutar
    organizer = DocumentOrganizer(config)

    try:
        organizer.organize_documents(origen_path, destino_path)
    except KeyboardInterrupt:
        print("\nProceso interrumpido por el usuario")
        sys.exit(1)
    except Exception as e:
        print(f"Error durante la ejecución: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
